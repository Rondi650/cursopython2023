# openpyxl - manipulando as planilhas do Workbook
# Com essa biblioteca será possível ler e escrever dados em células
# específicas, formatar células, inserir gráficos,
# criar fórmulas, adicionar imagens e outros elementos gráficos às suas
# planilhas. Ela é útil para automatizar tarefas envolvendo planilhas do
# Excel, como a criação de relatórios e análise de dados e/ou facilitando a
# manipulação de grandes quantidades de informações.
# Instalação necessária: pip install openpyxl
# Documentação: https://openpyxl.readthedocs.io/en/stable/
from pathlib import Path

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

wb = Workbook()

sheet_name = 'Minha planilha'
wb.create_sheet(sheet_name, 0)
ws: Worksheet = wb[sheet_name]
print(wb.worksheets)

wb.remove(wb['Sheet'])

ws.cell(1, 1, 'Nome')
ws.cell(1, 2, 'Idade')
ws.cell(1, 3, 'Nota')

students = [
    ['João', 14, 5.5],
    ['Maria', 13, 9.7],
    ['Luiz', 15, 8.8],
    ['Alberto', 16, 10],
    ['Rondi', 35, 100],
]

for i, student_row in enumerate(students, start=2): #linha
    for j, student_column in enumerate(student_row, start=1): #coluna
        ws.cell(i, j, student_column)

for student in students:
    ws.append(student)

wb.save(WORKBOOK_PATH)
